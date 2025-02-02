
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ...models.ore_productions_model import OreProductions
from ...models.ore_production_model import OreProductionsView
from ...models.mine_productions_model import mineProductionsView
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Sum
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from ...utils.permissions import get_dynamic_permissions

def format_angka(jumlah):
    if jumlah >= 1_000_000_000:
        return f"{jumlah / 1_000_000_000:.2f} B"
    elif jumlah >= 1_000_000:
        return f"{jumlah / 1_000_000:.2f} M"
    elif jumlah >= 1_000:
        return f"{jumlah / 1_000:.2f} K"
    else:
        return str(jumlah)

class viewMineProduction(View):

    def post(self, request):
        data_mine = self._datatables(request)
        return JsonResponse(data_mine, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = mineProductionsView.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(loader__icontains=search) |
                Q(hauler__icontains=search) |
                Q(hauler_class__icontains=search) |
                Q(sources_area__icontains=search) |
                Q(dumping_point__icontains=search) |
                Q(category_mine__icontains=search) |
                Q(nama_material__icontains=search) 
            )
       
        # Filter berdasarkan parameter dari request
        startDate       = request.POST.get('startDate')
        endDate         = request.POST.get('endDate')
        material_filter = request.POST.get('material_filter')
        sources_area    = request.POST.get('sources_area')
        loading_point   = request.POST.get('loading_point')
        dumping_point   = request.POST.get('dumping_point')
        dome_id         = request.POST.get('dome_id')
        category_mine   = request.POST.get('category_mine')

        if startDate and endDate:
            data = data.filter(date_production__range=[startDate, endDate])

        if material_filter:
            data = data.filter(nama_material=material_filter)

        if sources_area:
            data = data.filter(sources_area=sources_area)

        if loading_point:
            data = data.filter(loading_point=loading_point)

        if dumping_point:
            data = data.filter(dumping_point=dumping_point)

        if dome_id:
            data = data.filter(dome_id=dome_id)

        if category_mine:
            data = data.filter(category_mine=category_mine)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"             : item.id,
                "date_production": item.date_production,
                "shift"          : item.shift,
                "loader"         : item.loader,
                "hauler"         : item.hauler,
                "hauler_class"   : item.hauler_class,
                "sources_area"   : item.sources_area,
                "loading_point"  : item.loading_point,
                "dumping_point"  : item.dumping_point,
                "dome_id"        : item.dome_id,
                "category_mine"  : item.category_mine,
                "time_loading"   : item.time_loading,
                "time_dumping"   : item.time_dumping,
                "nama_material"  : item.nama_material,
                "ritase"         : item.ritase,
                "bcm"            : item.bcm,
                "tonnage"        : item.tonnage,
                "vendors"        : item.vendors,
                "mine_block"     : item.mine_block,
                "rl"             : item.rl,
                "remarks"        : item.remarks
                
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required   
@csrf_exempt
def export_mine_data(request):
    # Lakukan filter data sesuai parameter yang diterima dari permintaan
    startDate = request.GET.get('startDate')
    endDate = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    batch_status = request.GET.get('batch_status')
    area_filter = request.GET.get('area_filter')
    point_filter = request.GET.get('point_filter')
    source_filter = request.GET.get('source_filter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data Ore'

    # Write header row
    header = [
        'No', 
        'Date', 
        'Shift', 
        'Source',
        'From',
        'To-Rl',
        'Material',
        'Class',
        'Ni-Expect',
        'Mine Gelology',
        'Units',
        'Stockpile',
        'Dome',
        'Bacth',
        'Increment',
        'Batch Status',
        'Ritase',
        'Tonnage',
        'Dome Status',
        'Truck Factor',
        'Sample Id',
        'Remarks'
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'tgl_production', 
        'shift', 
        'prospect_area',
        'from_rl',
        'to_rl',
        'nama_material',
        'ore_class',
        'ni_grade',
        'grade_control',
        'unit_truck',
        'stockpile',
        'pile_id',
        'batch_code',
        'increment',
        'batch_status',
        'ritase',
        'tonnage',
        'pile_status',
        'truck_factor',
        'sample_number',
        'remarks',
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = OreProductionsView.objects.all().values_list(*columns)
    

    if startDate and endDate:
        queryset = queryset.filter(tgl_production__range=[startDate, endDate])
    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)

    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Ore_data.xlsx"'
    workbook.save(response)

    return response

@login_required
def total_mine_pds(request):

    queryset = mineProductionsView.objects.all()

    start_date      = request.GET.get('startDate')
    end_date        = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    sources_filter  = request.GET.get('sources_filter')
    loading_filter  = request.GET.get('loading_filter')
    dumping_filter  = request.GET.get('dumping_filter')
    dome_filter     = request.GET.get('dome_filter')

    if start_date and end_date:
        # start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        # end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
        queryset = queryset.filter(date_production__range=[start_date, end_date])

    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if sources_filter:
        queryset = queryset.filter(sources_area=sources_filter)
    if loading_filter:
        queryset = queryset.filter(loading_point=loading_filter)
    if dumping_filter:
        queryset = queryset.filter(dumping_point=dumping_filter)
    if dome_filter:
        queryset = queryset.filter(dome_id=dome_filter) 
        
    result = queryset.aggregate(
        qty     = Count('*'),
        bcm     = Sum('bcm', default=0),
        tonnage = Sum('tonnage', default=0)
    )

    return JsonResponse({
        'Qty'    : result['qty'],
        'Bcm'    : result['bcm'],
        'Tonnage': result['tonnage']
    })

@login_required
def total_pds_mining(request):
    queryset = mineProductionsView.objects.filter(category_mine='Mining')

    start_date      = request.GET.get('startDate')
    end_date        = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    sources_filter  = request.GET.get('sources_filter')
    loading_filter  = request.GET.get('loading_filter')
    dumping_filter  = request.GET.get('dumping_filter')
    dome_filter     = request.GET.get('dome_filter')

    if start_date and end_date:
        # start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        # end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
       queryset = queryset.filter(date_production__range=[start_date, end_date])


    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if sources_filter:
        queryset = queryset.filter(sources_area=sources_filter)
    if loading_filter:
        queryset = queryset.filter(loading_point=loading_filter)
    if dumping_filter:
        queryset = queryset.filter(dumping_point=dumping_filter)
    if dome_filter:
        queryset = queryset.filter(dome_id=dome_filter) 
        
    result = queryset.aggregate(
        qty     = Count('*'),
        bcm     = Sum('bcm', default=0),
        tonnage = Sum('tonnage', default=0)
    )

    return JsonResponse({
        'Qty'    : result['qty'],
        'Bcm'    : result['bcm'],
        'Tonnage': result['tonnage']
    })

@login_required
def total_pds_project(request):
    queryset = mineProductionsView.objects.filter(category_mine='Project')

    start_date      = request.GET.get('startDate')
    end_date        = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    sources_filter  = request.GET.get('sources_filter')
    loading_filter  = request.GET.get('loading_filter')
    dumping_filter  = request.GET.get('dumping_filter')
    dome_filter     = request.GET.get('dome_filter')

    if start_date and end_date:
        # start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        # end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
        queryset = queryset.filter(date_production__range=[start_date, end_date])


    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if sources_filter:
        queryset = queryset.filter(sources_area=sources_filter)
    if loading_filter:
        queryset = queryset.filter(loading_point=loading_filter)
    if dumping_filter:
        queryset = queryset.filter(dumping_point=dumping_filter)
    if dome_filter:
        queryset = queryset.filter(dome_id=dome_filter) 
        
    result = queryset.aggregate(
        qty     = Count('*'),
        bcm     = Sum('bcm', default=0),
        tonnage = Sum('tonnage', default=0)
    )

    return JsonResponse({
        'Qty'    : result['qty'],
        'Bcm'    : result['bcm'],
        'Tonnage': result['tonnage']
    })

@login_required
def getIdMinePds(request, id):
    if request.method == 'GET':
        try:
            items = OreProductions.objects.get(id=id)
            data = {
                'id'              : items.id,
                'tgl_production'  : items.tgl_production, 
                'shift'           : items.shift,
                'id_prospect_area': items.id_prospect_area,
                'id_block'        : items.id_block,
                'from_rl'         : items.from_rl,
                'to_rl'           : items.to_rl,
                'id_material'     : items.id_material,
                'grade_expect'    : items.grade_expect,
                'grade_control'   : items.grade_control,
                'unit_truck'      : items.unit_truck,
                'id_stockpile'    : items.id_stockpile,
                'id_pile'         : items.id_pile,
                'batch_code'      : items.batch_code,
                'increment'       : items.increment,
                'batch_status'    : items.batch_status,
                'ritase'          : items.ritase,
                'tonnage'         : items.tonnage,
                'pile_status'     : items.pile_status,
                'ore_class'       : items.ore_class,
                'truck_factor'    : items.truck_factor,
                'no_production'   : items.no_production,
                'sale_adjust'     : items.sale_adjust,
                'remarks'         : items.remarks
            }
            return JsonResponse(data)
        except OreProductions.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def delete_data_mine(request):
    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if job_id:
            # Lakukan penghapusan berdasarkan ID di sini
            data = OreProductions.objects.get(id=int(job_id))
            data.delete()
            return JsonResponse({'status': 'deleted'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
def mine_production_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    permissions = get_dynamic_permissions(request.user)
    context = {
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
          'permissions'   : permissions,
    }
    return render(request, 'admin-mine/list-productions.html',context)