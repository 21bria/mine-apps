from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.http import HttpResponse
from ...models.source_model import SourceMinesDumping
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.db import IntegrityError
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from ...utils.utils import clean_string
from ...utils.permissions import get_dynamic_permissions

@login_required
def minesDumping_page(request):
    permissions = get_dynamic_permissions(request.user)
    context = {
        'permissions'   : permissions,
    }
    return render(request, 'admin-mgoqa/master/list-source-dumping-point.html',context)

class sourceMinesDumping_List(View):

    def post(self, request):
        # Ambil semua data invoice yang valid
        MinesDumping = self._datatables(request)
        return JsonResponse(MinesDumping, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Set record total
        records_total = SourceMinesDumping.objects.all().count()
        # Set records filtered
        records_filtered = records_total
        # Ambil semua yang valid
        data = SourceMinesDumping.objects.all()

        if search:
            data = SourceMinesDumping.objects.filter(
                Q(dumping_point__icontains=search) |
                Q(remarks__icontains=search)|
                Q(category__icontains=search)
            )
            records_total = data.count()
            records_filtered = records_total

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Atur paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"            : item.id,
                "dumping_point" : item.dumping_point,
                "remarks"       : item.remarks,
                "category"      : item.category,
                "status"        : item.status,
            } for item in object_list
        ]

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

@login_required        
@csrf_exempt
def get_minesDumping(request, id):
    allowed_groups = ['superadmin', 'admin-mining','admin-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )

    if request.method == 'GET':
        try:
            job = SourceMinesDumping.objects.get(id=id)
            data = {
                'id'            : job.id,
                'dumping_point' : clean_string(job.dumping_point), 
                'remarks'       : clean_string(job.remarks),
                'category'      : clean_string(job.category),
                'status'        : clean_string(job.status),
                'created_at'    : job.created_at
            }
            return JsonResponse(data)
        except SourceMinesDumping.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required   
def insert_minesDumping(request):
    allowed_groups = ['superadmin', 'admin-mining','admin-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    if request.method == 'POST':
        dumping_point = request.POST.get('dumping_point')
        remarks       = request.POST.get('remarks')
        category      = request.POST.get('category')
        status        = 1

        try:
            new_job = SourceMinesDumping.objects.create(
                    dumping_point=dumping_point,
                    remarks=remarks,
                    category =category,
                    status=status
                    )
            return JsonResponse({
                'status' : 'success',
                'message': 'Data berhasil disimpan.',
                'data': {
                    'id'            : new_job.id,
                    'dumping_point' : new_job.dumping_point,
                    'remarks'       : new_job.remarks,
                    'category'      : new_job.category,
                    'status'        : new_job.status,
                    'created_at'    : new_job.created_at
                }
            })
        except IntegrityError as e:
            # Check if the error is a duplicate entry error
            if 'Duplicate entry' in str(e):
                return JsonResponse({'status': 'error', 'message': 'The data already exists'}, status=403)
            else:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Metode tidak diizinkan'}, status=405)

@login_required
def update_minesDumping(request, id):
    allowed_groups = ['superadmin', 'admin-mining','admin-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    if request.method == 'POST':
        try:
            job = SourceMinesDumping.objects.get(id=id)
            job.dumping_point  = request.POST.get('dumping_point')
            job.remarks     = request.POST.get('remarks')
            job.category      = request.POST.get('category')
            job.save()

            return JsonResponse({
                'id'            : job.id,
                'dumping_point' : job.dumping_point,
                'remarks'       : job.remarks,
                'category'      : job.category,
                'created_at'    : job.created_at,
                'updated_at'    : job.updated_at
            })
        
        except SourceMinesDumping.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)
        except IntegrityError as e:
            error_message = str(e)
            if 'Duplicate entry' in error_message:
                 return JsonResponse({'error': 'Duplikat data: The data already exists', 'message': error_message}, status=400)
            else:
                 return JsonResponse({'error': 'Error pada operasi database', 'message': error_message}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Metode tidak diizinkan'}, status=405)

@login_required
def delete_minesDumping(request):
    allowed_groups = ['superadmin']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if job_id:
            data = SourceMinesDumping.objects.get(id=int(job_id))
            data.delete()
            return JsonResponse({'status': 'deleted'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


@csrf_exempt
def export_dumping_point(request):
    sourceFilter   = request.POST.get('sourceFilter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Dumping Point'

    # Write header row
    header = [
        'No', 
        'Name', 
        'Remarks', 
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'dumping_point', 
        'remarks'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = SourceMinesDumping.objects.all().values_list(*columns)

    # Membersihkan setiap nilai string dalam hasil queryset
    cleaned_data = [
        [clean_string(value) for value in row]  # Terapkan clean_string ke setiap nilai dalam baris
        for row in queryset
    ]
    
    # if sourceFilter:
    #     queryset = queryset.filter(sampling_area=sourceFilter)

    for row_num, (row_count, row) in enumerate(enumerate(cleaned_data, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            
            if isinstance(cell_value, str):  
                try:
                    cell_value = float(cell_value.replace(',', '').replace(' ', '').replace('.', '', 1))
                except ValueError:
                    pass  # Biarkan tetap string jika bukan angka

            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dumping_point.xlsx"'
    workbook.save(response)

    return response