from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ...models.selling_data_model import SellingProductions
from ...models.selling_details_view_model import SellingDetailsView
from ...models.source_model import SourceMinesDumping,SourceMinesDome
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.utils import get_column_letter
from ...utils.permissions import get_dynamic_permissions

@login_required
def sale_details_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    permissions = get_dynamic_permissions(request.user)
   
    context = {
        'permissions' : permissions,
        'start_date'  : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'    : today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-mgoqa/selling/list-selling-details.html',context)

class SellingDetails(View):
    def post(self, request):
        # Ambil semua data invoice yang valid
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = SellingDetailsView.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(sampling_area__icontains=search) |
                Q(sampling_point__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(delivery_order__icontains=search) |
                Q(haulage_code__icontains=search) |
                Q(factory_stock__icontains=search)
            )
       

        # Filter berdasarkan parameter dari request
        from_date       = request.POST.get('from_date')
        to_date         = request.POST.get('to_date')
        materialFilter  = request.POST.get('materialFilter')
        areaFilter      = request.POST.get('areaFilter')
        pointFilter     = request.POST.get('pointFilter')
        factoriesFilter = request.POST.get('factoriesFilter')
        productFilter   = request.POST.get('productFilter')

        if from_date and to_date:
            data = data.filter(timbang_isi__range=[from_date, to_date])

        if materialFilter:
            data = data.filter(nama_material=materialFilter)

        if areaFilter:
            data = data.filter(sampling_area=areaFilter)

        if pointFilter:
            data = data.filter(sampling_point=pointFilter)

        if factoriesFilter:
            data = data.filter(factory_stock=factoriesFilter)

        if productFilter:
            data = data.filter(delivery_order=productFilter)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"             : item.id,
                "tgl_hauling"    : item.tgl_hauling,
                "shift"          : item.shift,
                "sampling_area"  : item.sampling_area,
                "sampling_point" : item.sampling_point,
                "nama_material"  : item.nama_material,
                "batch"          : item.batch,
                "new_scci_sub"   : item.new_scci_sub,
                "new_awk_sub"    : item.new_awk_sub,
                "date_wb"        : item.date_wb,
                "fill_weigth_f"  : item.fill_weigth_f,
                "empety_weigth_f": item.empety_weigth_f,
                "netto_kg"       : item.netto_kg,
                "netto_ton"      : item.netto_ton,
                "tonnage"        : item.tonnage,
                "delivery_order" : item.delivery_order,
                "factory_stock"  : item.factory_stock,
                "haulage_code"   : item.haulage_code,
                "tonnage"        : item.tonnage
                
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required    
@csrf_exempt       
def export_sale_data(request):
    # Lakukan filter data sesuai parameter yang diterima dari permintaan
    from_date       = request.POST.get('from_date')
    to_date         = request.POST.get('to_date')
    materialFilter  = request.POST.get('materialFilter')
    areaFilter      = request.POST.get('areaFilter')
    pointFilter     = request.POST.get('pointFilter')
    factoriesFilter = request.POST.get('factoriesFilter')
    productFilter   = request.POST.get('productFilter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data Ore'

    # Write header row
    header = [
        'No', 
        'Date Hauling',
        'Week',
        'Month',
        'Year',
        'Shift',
        'Stockpile',
        'PileID',
        'Materials',
        'Truck',
        'Batch',
        'SCCI',
        'AWK',
        'Date Weigth',
        'Berat[Isi]/Kg',
        'Berat[Kosong]/Kg',
        'Netto/Kg',
        'Tonnage',
        'Factory',
        'Product Order'
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'tgl_hauling', 
        'minggu', 
        'bulan',
        'tahun',
        'shift',
        'sampling_area',
        'sampling_point',
        'nama_material',
        'unit_code',
        'batch',
        'new_scci_sub',
        'new_awk_sub',
        'timbang_isi',
        'fill_weigth_f',
        'empety_weigth_f',
        'netto_kg',
        'netto_ton',
        'factory_stock',
        'delivery_order'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = SellingDetailsView.objects.all().values_list(*columns)
    
    if from_date and to_date:
            queryset = queryset.filter(timbang_isi__range=[from_date, to_date])

    if materialFilter:
            queryset = queryset.filter(nama_material=materialFilter)

    if areaFilter:
            queryset = queryset.filter(sampling_area=areaFilter)

    if pointFilter:
            queryset = queryset.filter(sampling_point=pointFilter)

    if factoriesFilter:
            queryset = queryset.filter(factory_stock=factoriesFilter)

    if productFilter:
            queryset = queryset.filter(delivery_order=productFilter)

    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Selling_data.xlsx"'
    workbook.save(response)

    return response

@login_required
def getIdSale(request, id):
    if request.method == 'GET':
        try:
            items = SellingProductions.objects.get(id=id)
            dumping_point = None
            dome_point    = None
            
            if items.id_stockpile:
                dumping = SourceMinesDumping.objects.filter(id=items.id_stockpile).first()
                if dumping:
                    dumping_point = dumping.dumping_point

            if items.id_pile:
                dome = SourceMinesDome.objects.filter(id=items.id_pile).first()
                if dome:
                    dome_point = dome.pile_id

            data = {
                'id'                : items.id,
                'tgl_hauling'       : items.tgl_hauling, 
                'shift'             : items.shift,
                'id_material'       : items.id_material,
                'id_stockpile'      : items.id_stockpile,
                'id_pile'           : items.id_pile,
                'dumping_point'     : dumping_point,
                'dome_point'        : dome_point,
                'id_truck'          : items.id_truck,
                'delivery_order'    : items.delivery_order,
                'nota'              : items.nota,
                'id_factory'        : items.id_factory,
                'type_selling'      : items.type_selling,
                'empety_weigth_f'   : items.empety_weigth_f,
                'fill_weigth_f'     : items.fill_weigth_f,
                'netto_weigth_f'    : items.netto_weigth_f,
                'id_stock_temp'     : items.id_stock_temp,
                'id_dome_temp'      : items.id_dome_temp,
                'no_input'          : items.no_input,
                'remarks'           : items.remarks,
                'batch_g'           : items.batch_g,
                'kode_batch_g'      : items.kode_batch_g,
                'new_scci'          : items.new_scci,
                'new_scci_sub'      : items.new_scci_sub,
                'new_kode_batch_scci': items.new_kode_batch_scci,
                'new_awk'           : items.new_awk,
                'new_awk_sub'       : items.new_awk_sub,
                'new_kode_batch_awk': items.new_kode_batch_awk,
                'new_batch_awk_pulp': items.new_batch_awk_pulp,
                'awk_order'         : items.awk_order,
                'scci_order'        : items.scci_order,
                'load_code'         : items.load_code,
                'haulage_code'      : items.haulage_code,
                'date_wb'           : items.date_wb.strftime('%Y-%m-%d %H:%M:%S'),
                'timbang_kosong'    : items.timbang_kosong.strftime('%Y-%m-%d %H:%M:%S'),
                'timbang_isi'       : items.timbang_isi.strftime('%Y-%m-%d %H:%M:%S'),
                'sale_adjust'       : items.sale_adjust,
                'sale_dome'         : items.sale_dome,
            }
            return JsonResponse(data)
        except SellingProductions.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def delete_sale(request):
    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if job_id:
            # Lakukan penghapusan berdasarkan ID di sini
            data = SellingProductions.objects.get(id=int(job_id))
            data.delete()
            return JsonResponse({'status': 'deleted'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
