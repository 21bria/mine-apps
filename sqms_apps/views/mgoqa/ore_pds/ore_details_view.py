from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.details_mral_view_model import DetailsMral
from ....models.details_roa_view_model import DetailsRoa
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from ....utils.permissions import get_dynamic_permissions

@login_required
def ore_details_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    # Cek permission
    permissions = get_dynamic_permissions(request.user)
    context = {
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
        'permissions': permissions,
    }
    return render(request, 'admin-mgoqa/production-ore/list-ore-details.html',context)


class OreDetailsView(View):

    def post(self, request):
        # Ambil semua data invoice yang valid
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = DetailsMral.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(prospect_area__icontains=search) |
                Q(mine_block__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(ore_class__icontains=search) |
                Q(grade_control__icontains=search) |
                Q(unit_truck__icontains=search) |
                Q(stockpile__icontains=search) |
                Q(pile_id__icontains=search) |
                Q(batch_code__icontains=search) |
                Q(batch_status__icontains=search) |
                Q(truck_factor__icontains=search) |
                Q(sample_number__icontains=search) 
            )
       

        # Filter berdasarkan parameter dari request
        startDate       = request.POST.get('startDate')
        endDate         = request.POST.get('endDate')
        material_filter = request.POST.get('material_filter')
        batch_status    = request.POST.get('batch_status')
        area_filter     = request.POST.get('area_filter')
        point_filter    = request.POST.get('point_filter')
        source_filter   = request.POST.get('source_filter')

        if startDate and endDate:
            data = data.filter(tgl_production__range=[startDate, endDate])

        if material_filter:
            data = data.filter(nama_material=material_filter)

        if batch_status:
            data = data.filter(batch_status=batch_status)

        if area_filter:
            data = data.filter(stockpile=area_filter)

        if point_filter:
            data = data.filter(pile_id=point_filter)

        if source_filter:
            data = data.filter(prospect_area=source_filter)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [

            {
                "tgl_production": item.tgl_production,
                "category": item.category,
                "shift": item.shift,
                "prospect_area": item.prospect_area,
                "mine_block": item.mine_block,
                "rl": (str(item.from_rl) if item.from_rl is not None else '') + '-' + (str(item.to_rl) if item.to_rl is not None else ''),
                # "to_rl": ,
                "nama_material": item.nama_material,
                "ore_class": item.ore_class,
                "ni_grade": item.ni_grade,
                "grade_control": item.grade_control,
                "unit_truck": item.unit_truck,
                "stockpile": item.stockpile,
                "pile_id": item.pile_id,
                "batch_code": item.batch_code,
                "increment": item.increment,
                "batch_status": item.batch_status,
                "ritase": item.ritase,
                "tonnage": item.tonnage,
                "pile_status": item.pile_status,
                "truck_factor": item.truck_factor,
                "remarks": item.remarks,
                "sample_number": item.sample_number,
                "MRAL_Ni": item.MRAL_Ni
                
            } for item in object_list
        ]

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': total_records_filtered,
            'data': data,
            'start': start,
            'length': length,
            'totalPages': total_pages,
        }
    
@csrf_exempt       
def export_detail_mral(request):
    # Lakukan filter data sesuai parameter yang diterima dari permintaan
    startDate = request.GET.get('startDate')
    endDate = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    batch_status = request.GET.get('batch_status')
    area_filter = request.GET.get('area_filter')
    point_filter = request.GET.get('point_filter')
    source_filter = request.GET.get('source_filter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data Ore - mral'

    # Write header row
    header = [
        'No', 
        'Date', 
        'Shift', 
        'Source',
        'From',
        'To-Rl',
        'Material',
        'Class',
        'Ni-Expect',
        'Mine Gelology',
        'Units',
        'Stockpile',
        'Dome',
        'Bacth',
        'Increment',
        'Batch Status',
        'Ritase',
        'Tonnage',
        'Dome Status',
        'Truck Factor',
        'Sample Id',
        'Remarks',
        'Ni',
        'Co',
        'Fe2O3  ',
        'Fe',
        'MgO',
        'SiO2',
        'Sm'
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'tgl_production', 
        'shift', 
        'prospect_area',
        'from_rl',
        'to_rl',
        'nama_material',
        'ore_class',
        'ni_grade',
        'grade_control',
        'unit_truck',
        'stockpile',
        'pile_id',
        'batch_code',
        'increment',
        'batch_status',
        'ritase',
        'tonnage',
        'pile_status',
        'truck_factor',
        'sample_number',
        'remarks',
        'MRAL_Ni',
        'MRAL_Co',
        'MRAL_Fe2O3',
        'MRAL_Fe',
        'MRAL_MgO',
        'MRAL_SiO2',
        'MRAL_SM'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = DetailsMral.objects.all().values_list(*columns)
    

    if startDate and endDate:
        queryset = queryset.filter(tgl_production__range=[startDate, endDate])
    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)

    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Ore_data_mral.xlsx"'
    workbook.save(response)

    return response

def export_detail_roa(request):
    # Lakukan filter data sesuai parameter yang diterima dari permintaan
    startDate = request.GET.get('startDate')
    endDate = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    batch_status = request.GET.get('batch_status')
    area_filter = request.GET.get('area_filter')
    point_filter = request.GET.get('point_filter')
    source_filter = request.GET.get('source_filter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data Ore - Roa'

    # Write header row
    header = [
        'No', 
        'Date', 
        'Shift', 
        'Source',
        'From',
        'To-Rl',
        'Material',
        'Class',
        'Ni-Expect',
        'Mine Gelology',
        'Units',
        'Stockpile',
        'Dome',
        'Bacth',
        'Increment',
        'Batch Status',
        'Ritase',
        'Tonnage',
        'Dome Status',
        'Sample Id',
        'Remarks',
        'Ni',
        'Co',
        'Al2O3',
        'CaO',
        'Cr2O3',
        'Fe2O3',
        'Fe',
        'MgO',
        'SiO2',
        'Sm',
        'Mc',
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'tgl_production', 
        'shift', 
        'prospect_area',
        'from_rl',
        'to_rl',
        'nama_material',
        'ore_class',
        'ni_grade',
        'grade_control',
        'unit_truck',
        'stockpile',
        'pile_id',
        'batch_code',
        'increment',
        'batch_status',
        'ritase',
        'tonnage',
        'pile_status',
        'sample_number',
        'remarks',
        'ROA_Ni',
        'ROA_Co',
        'ROA_Al2O3',
        'ROA_CaO',
        'ROA_Cr2O3',
        'ROA_Fe2O3',
        'ROA_Fe',
        'ROA_MgO',
        'ROA_SiO2',
        'ROA_SM',
        'ROA_MC'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = DetailsRoa.objects.all().values_list(*columns)
    

    if startDate and endDate:
        queryset = queryset.filter(tgl_production__range=[startDate, endDate])
    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)

    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Ore_data_roa.xlsx"'
    workbook.save(response)

    return response

