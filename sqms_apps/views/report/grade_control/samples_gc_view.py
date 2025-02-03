from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.http import HttpResponse
from ....models.sample_grade_control_model import GradeControlSamples
from django.db.models import Q
from django.shortcuts import render
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views import View
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from ....utils.utils import clean_string
from ....utils.permissions import get_dynamic_permissions
import locale

# Atur locale untuk menangani format angka dengan koma sebagai pemisah ribuan
locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8')

@login_required
def sample_gc_page(request):
    allowed_groups = ['superadmin','admin-mgoqa','superintendent-mgoqa','manager-mgoqa','user-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        # Jika tidak memiliki izin, arahkan ke halaman error
        context = {
            'error_message': 'You do not have permission to access this page.',
        }
        return render(request, '403.html', context, status=403)
    
    # Dapatkan tanggal hari ini
    today = datetime.today()

    # Hitung tanggal awal minggu (Senin)
    start_of_week = today - timedelta(days=today.weekday())

    # Hitung tanggal akhir minggu (Minggu)
    end_of_week = start_of_week + timedelta(days=6)

    permissions = get_dynamic_permissions(request.user)

    # Menambahkan tanggal awal dan akhir minggu ke konteks
    context = {
        'start_of_week': start_of_week.strftime('%Y-%m-%d'),
        'end_of_week'  : end_of_week.strftime('%Y-%m-%d'),
        'permissions'  : permissions,
    }

    return render(request, 'admin-mgoqa/report-gc/list-sample-grade-control.html', context)

class GcSamples(View):
    def post(self, request):
        data_sample = self._datatables(request)
        return JsonResponse(data_sample, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = GradeControlSamples.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(sampling_area__icontains=search) |
                Q(type_sample__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(sample_method__icontains=search) |
                Q(sample_id__icontains=search) |
                Q(job_number__icontains=search)
            )
       

        # Filter berdasarkan parameter dari request
        from_date      = request.POST.get('from_date')
        to_date        = request.POST.get('to_date')
        materialFilter = request.POST.get('materialFilter')
        method_filter  = request.POST.get('method_filter')
        sourceFilter   = request.POST.get('sourceFilter')


        if from_date and to_date:
            data = data.filter(tgl_sample__range=[from_date, to_date])

        if method_filter:
            data = data.filter(sample_method=method_filter)

        if materialFilter:
            data = data.filter(nama_material=materialFilter)

        if sourceFilter:
            data = data.filter(sampling_area=sourceFilter)


        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        # Format data sesuai dengan DataTables
        data = [
                {
                    "tgl_sample"   : item.tgl_sample,
                    "tgl_deliver"  : item.tgl_deliver,
                    "sampling_area": item.sampling_area,
                    "sample_id"    : item.sample_id,
                    "sample_method": item.sample_method,
                    "nama_material": item.nama_material,
                    "job_number"   : item.job_number,
                    "release_date" : item.release_date,
                    "ni"           : item.ni,
                    "co"           : item.co,
                    "fe2o3"        : item.fe2o3,
                    "fe"           : item.fe,
                    "mgo"          : item.mgo,
                    "sio2"         : item.sio2,
                    "sm"           : item.sm
                } for item in object_list
            ]

        # print(data)

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@csrf_exempt
def export_gc_sample(request):
    from_date      = request.POST.get('from_date')
    to_date        = request.POST.get('to_date')
    materialFilter = request.POST.get('materialFilter')
    method_filter  = request.POST.get('method_filter')
    sourceFilter   = request.POST.get('sourceFilter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data GC Samples'

    # Write header row
    header = [
        'No', 
        'Date sample', 
        'Deliver', 
        'Sampling area',
        'Sample id',
        'Type sample',
        'Sample method',
        'Material',
        'Job number',
        'Release date',
        'ni',
        'co',
        'fe2o3',
        'fe',
        'mgo',
        'sio2',
        # 'sm'
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'tgl_sample', 
        'tgl_deliver', 
        'sampling_area',
        'sample_id',
        'type_sample',
        'sample_method',
        'nama_material',
        'job_number',
        'release_date',
        'ni',
        'co',
        'fe2o3',
        'fe',
        'mgo',
        'sio2',
        # 'sm'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = GradeControlSamples.objects.all().values_list(*columns)
    
    if from_date and to_date:
        queryset = queryset.filter(tgl_sample__range=[from_date, to_date])

    if method_filter:
        queryset = queryset.filter(sample_method=method_filter)

    if materialFilter:
        queryset = queryset.filter(nama_material=materialFilter)

    if sourceFilter:
        queryset = queryset.filter(sampling_area=sourceFilter)

    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            
            # Coba konversi angka yang memiliki koma sebagai pemisah ribuan
            if isinstance(cell_value, str):  
                try:
                    # Hapus koma sebelum konversi dan pastikan format desimal benar
                    cell_value = float(cell_value.replace(',', '').replace(' ', '').replace('.', '', 1))  # Hapus koma ribuan dan titik desimal hanya satu kali
                except ValueError:
                    pass  # Biarkan tetap string jika bukan angka

            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_gc.xlsx"'
    workbook.save(response)

    return response
