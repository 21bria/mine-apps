from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.http import JsonResponse
from django.http import HttpResponse
from django.db import connections
from datetime import datetime, timedelta
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import numpy as np  # Import numpy for NaN representation
import plotly.graph_objs as go
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from .....utils.permissions import get_dynamic_permissions

@login_required
def sampleCrmMralPage(request):
    allowed_groups = ['superadmin','admin-mgoqa','superintendent-mgoqa','manager-mgoqa','user-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        # Jika tidak memiliki izin, arahkan ke halaman error
        context = {
            'error_message': 'You do not have permission to access this page.',
        }
        return render(request, '403.html', context, status=403)
    today = datetime.today()
    last_monday = today - timedelta(days=today.weekday())

    permissions = get_dynamic_permissions(request.user)

    context = {
        'start_date': last_monday.strftime('%Y-%m-%d'),
        'end_date': today.strftime('%Y-%m-%d'),
        'permissions': permissions,
    }
    # form = DateFilterForm(request.GET or None)
    return render(request, 'admin-mgoqa/report-qa/list-samples-crm-mral.html',context)

@login_required
def sampleCrmMralChart(request):
    allowed_groups = ['superadmin','admin-mgoqa','superintendent-mgoqa','manager-mgoqa','user-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        # Jika tidak memiliki izin, arahkan ke halaman error
        context = {
            'error_message': 'You do not have permission to access this page.',
        }
        return render(request, '403.html', context, status=403)
    today = datetime.today()
    last_monday = today - timedelta(days=today.weekday())

    permissions = get_dynamic_permissions(request.user)
    context = {
        'start_date': last_monday.strftime('%Y-%m-%d'),
        'end_date': today.strftime('%Y-%m-%d'),
        'permissions': permissions,
    }
    # form = DateFilterForm(request.GET or None)
    return render(request, 'admin-mgoqa/report-qa/chart-sample-crm-mral.html',context)

@login_required
def getDataCrmMral(request):
    # Mendapatkan nilai filter dari request
    startDate  = request.GET.get('start_date')
    endDate    = request.GET.get('end_date')
    TypeCrm    = request.GET.get('filterTypeCrm')

    # Query dengan parameterisasi untuk keamanan
    query = """
            SELECT 
                    oreas_name, oreas_ni, oreas_fe,
                    oreas_mgo, oreas_sio2,sample_number, 
                    sampling_deskripsi, 
                    sample_id, release_date, mral_ni, mral_fe, mral_mgo, mral_sio2,
                    CASE 
                        WHEN mral_ni < oreas_ni + (oreas_ni * -0.1) OR mral_ni > oreas_ni + (oreas_ni * 0.1) THEN 1 ELSE 0
                    END AS cek_ni,
                    CASE
                        WHEN mral_fe < oreas_fe + (oreas_fe * -0.1) OR mral_fe > oreas_fe + (oreas_fe * 0.1) THEN 1 ELSE 0
                    END AS cek_fe,
                    CASE
                        WHEN mral_mgo < oreas_mgo + (oreas_mgo * -0.1) OR mral_mgo > oreas_mgo + (oreas_mgo * 0.1) THEN 1 ELSE 0
                    END AS cek_mgo,
                    CASE
                        WHEN mral_sio2 < oreas_sio2 + (oreas_sio2 * -0.1) OR mral_sio2 > oreas_sio2 + (oreas_sio2 * 0.1) THEN 1 ELSE 0
                    END AS cek_sio2,
                        -- ni
                    round(oreas_ni+(oreas_ni*+0.1),2) plus_ni_10,
                    round(oreas_ni+(oreas_ni*+0.05),2) plus_ni_5,
                    round(oreas_ni+(oreas_ni*-0.05),2)min_ni_5,
                    round(oreas_ni+(oreas_ni*-0.1),2)min_ni_10,
                        -- fe
                    round(oreas_fe+(oreas_fe*+0.1),2) plus_fe_10,
                    round(oreas_fe+(oreas_fe*+0.05),2) plus_fe_5,
                    round(oreas_fe+(oreas_fe*-0.05),2)min_fe_5,
                    round(oreas_fe+(oreas_fe*-0.1),2)min_fe_10,
                        -- mgo
                    round(oreas_mgo+(oreas_mgo*+0.1),2) plus_mgo_10,
                    round(oreas_mgo+(oreas_mgo*+0.05),2) plus_mgo_5,
                    round(oreas_mgo+(oreas_mgo*-0.05),2)min_mgo_5,
                    round(oreas_mgo+(oreas_mgo*-0.1),2)min_mgo_10,
                        -- sio2
                    round( oreas_sio2+( oreas_sio2*+0.1),2) plus_sio2_10,
                    round( oreas_sio2+( oreas_sio2*+0.05),2) plus_sio2_5,
                    round( oreas_sio2+( oreas_sio2*-0.05),2)min_sio2_5,
                    round( oreas_sio2+( oreas_sio2*-0.1),2)min_sio2_10
            FROM oreas_diff_abs_mral  
            WHERE release_date >= %s AND release_date <= %s  AND oreas_name =%s
    """

    params = (startDate, endDate,TypeCrm)
    df = pd.read_sql_query(query, connections['sqms_db'], params=params)

    # Menghitung nilai dan average
    jumlah    = len(df)  # Menggunakan len(df) untuk menghitung jumlah baris
    true_ni   = len(df[df['cek_ni'] == 0])  # Count where 'cek_ni' column equals 0
    false_ni  = len(df[df['cek_ni'] == 1])  # Count where 'cek_ni' column equals 1
    true_fe   = len(df[df['cek_fe'] == 0])  
    false_fe  = len(df[df['cek_fe'] == 1])  
    true_mgo  = len(df[df['cek_mgo'] == 0]) 
    false_mgo = len(df[df['cek_mgo'] == 1]) 
    true_sio2 = len(df[df['cek_sio2'] == 0]) 
    false_sio2= len(df[df['cek_sio2'] == 1]) 
    avg_ni      = df['mral_ni'].mean(skipna=True)
    avg_fe      = df['mral_fe'].mean(skipna=True)
    avg_mgo     = df['mral_mgo'].mean(skipna=True)
    avg_sio2    = df['mral_sio2'].mean(skipna=True)



     # load data
    response_data = {
                'list_data'  : df.to_dict('records'),
                'jumlah'     : jumlah,
                'avg_ni'     : avg_ni if not np.isnan(avg_ni) else None,
                'avg_fe'     : avg_fe if not np.isnan(avg_fe) else None,
                'avg_mgo'    : avg_mgo if not np.isnan(avg_mgo) else None,
                'avg_sio2'   : avg_sio2 if not np.isnan(avg_sio2) else None,
                'true_ni'    :true_ni,
                'false_ni'   :false_ni,
                'true_fe'    :true_fe,
                'false_fe'   :false_fe,
                'true_mgo'   :true_mgo,
                'false_mgo'  :false_mgo,
                'true_sio2'  :true_sio2,
                'false_sio2' :false_sio2
    }

    return JsonResponse(response_data)

# plotly | graphing libaries
@login_required
def getDataCrmMralPloty(request):
    # Mendapatkan nilai filter dari request
    startDate = request.GET.get('start_date')
    endDate   = request.GET.get('end_date')
    TypeCrm   = request.GET.get('filterTypeCrm')

    # Query dengan parameterisasi untuk keamanan
    query = """
            SELECT 
                    oreas_name, oreas_ni, oreas_fe,
                    oreas_mgo, oreas_sio2, sample_number, 
                    sampling_deskripsi, 
                    sample_id, release_date, mral_ni, mral_fe, mral_mgo, mral_sio2,
                    CASE 
                        WHEN mral_ni < oreas_ni + (oreas_ni * -0.1) OR mral_ni > oreas_ni + (oreas_ni * 0.1) THEN 1 ELSE 0
                    END AS cek_ni,
                    CASE
                        WHEN mral_fe < oreas_fe + (oreas_fe * -0.1) OR mral_fe > oreas_fe + (oreas_fe * 0.1) THEN 1 ELSE 0
                    END AS cek_fe,
                    CASE
                        WHEN mral_mgo < oreas_mgo + (oreas_mgo * -0.1) OR mral_mgo > oreas_mgo + (oreas_mgo * 0.1) THEN 1 ELSE 0
                    END AS cek_mgo,
                    CASE
                        WHEN mral_sio2 < oreas_sio2 + (oreas_sio2 * -0.1) OR mral_sio2 > oreas_sio2 + (oreas_sio2 * 0.1) THEN 1 ELSE 0
                    END AS cek_sio2,
                    round(oreas_ni + (oreas_ni * +0.1), 2) as plus_ni_10,
                    round(oreas_ni + (oreas_ni * +0.05), 2) as plus_ni_5,
                    round(oreas_ni + (oreas_ni * -0.05), 2) as min_ni_5,
                    round(oreas_ni + (oreas_ni * -0.1), 2) as min_ni_10,
                    round(oreas_fe + (oreas_fe * +0.1), 2) as plus_fe_10,
                    round(oreas_fe + (oreas_fe * +0.05), 2) as plus_fe_5,
                    round(oreas_fe + (oreas_fe * -0.05), 2) as min_fe_5,
                    round(oreas_fe + (oreas_fe * -0.1), 2) as min_fe_10,
                    round(oreas_mgo + (oreas_mgo * +0.1), 2) as plus_mgo_10,
                    round(oreas_mgo + (oreas_mgo * +0.05), 2) as plus_mgo_5,
                    round(oreas_mgo + (oreas_mgo * -0.05), 2) as min_mgo_5,
                    round(oreas_mgo + (oreas_mgo * -0.1), 2) as min_mgo_10,
                    round(oreas_sio2 + (oreas_sio2 * +0.1), 2) as plus_sio2_10,
                    round(oreas_sio2 + (oreas_sio2 * +0.05), 2) as plus_sio2_5,
                    round(oreas_sio2 + (oreas_sio2 * -0.05), 2) as min_sio2_5,
                    round(oreas_sio2 + (oreas_sio2 * -0.1), 2) as min_sio2_10
            FROM oreas_diff_abs_mral  
            WHERE release_date >= %s AND release_date <= %s AND oreas_name = %s
    """

    params = (startDate, endDate, TypeCrm)
    df = pd.read_sql_query(query, connections['sqms_db'], params=params)
    
    # Jika data kosong 
    if df.empty:
        fig_ni   = go.Figure()
        fig_fe   = go.Figure()
        fig_mgo  = go.Figure()
        fig_sio2 = go.Figure()
        fig_ni.update_layout(
            xaxis={"visible": False},
            yaxis={"visible": False},
            plot_bgcolor='rgba(0,0,0,0)', 
            annotations=[
                {
                    "text": "No matching data found",
                    "xref": "paper",
                    "yref": "paper",
                    "showarrow": False,
                    "font": {"size": 14}
                }
            ],
            height=265,
        )
        fig_fe.update_layout(
            xaxis={"visible": False},
            yaxis={"visible": False},
            plot_bgcolor='rgba(0,0,0,0)', 
            annotations=[
                {
                    "text": "No matching data found",
                    "xref": "paper",
                    "yref": "paper",
                    "showarrow": False,
                    "font": {"size": 14}
                }
            ],
            height=265,
        )
        fig_mgo.update_layout(
            xaxis={"visible": False},
            yaxis={"visible": False},
            plot_bgcolor='rgba(0,0,0,0)', 
            annotations=[
                {
                    "text": "No matching data found",
                    "xref": "paper",
                    "yref": "paper",
                    "showarrow": False,
                    "font": {"size": 14}
                }
            ],
            height=265,
        )
        fig_sio2.update_layout(
            xaxis={"visible": False},
            yaxis={"visible": False},
            plot_bgcolor='rgba(0,0,0,0)', 
            annotations=[
                {
                    "text": "No matching data found",
                    "xref": "paper",
                    "yref": "paper",
                    "showarrow": False,
                    "font": {"size": 14}
                }
            ],
            height=265,
        )

        plot_ni   = fig_ni.to_html(full_html=False)
        plot_fe   = fig_fe.to_html(full_html=False)
        plot_mgo  = fig_mgo.to_html(full_html=False)
        plot_sio2 = fig_sio2.to_html(full_html=False)

        return JsonResponse({'plotNi':plot_ni,'plotFe':plot_fe,'plotMgo':plot_mgo,'plotMgo':plot_mgo,'plotSio2':plot_sio2})

    # Menghitung nilai dan average
    jml_row   = len(df)
    true_ni   = len(df[df['cek_ni'] == 0])
    false_ni  = len(df[df['cek_ni'] == 1])
    avg_ni    = df['mral_ni'].mean()
    
    true_fe   = len(df[df['cek_fe'] == 0])  
    false_fe  = len(df[df['cek_fe'] == 1]) 
    avg_fe    = df['mral_fe'].mean(skipna=True)

    true_mgo  = len(df[df['cek_mgo'] == 0]) 
    false_mgo = len(df[df['cek_mgo'] == 1]) 
    avg_sio2  = df['mral_sio2'].mean(skipna=True)

    true_sio2 = len(df[df['cek_sio2'] == 0]) 
    false_sio2= len(df[df['cek_sio2'] == 1]) 
    avg_mgo   = df['mral_mgo'].mean(skipna=True)
    

    # Warna untuk masing-masing trace
    colors = {
        'grade'      : '#EE2E31',
        'certificate': '#1d7874',
        'dil_10'     : '#071E22',
        'dil_5'      : '#57C5B6',
        'mean'       : '#FFB100'
    }

    # Create Plotly figure - Ni
    fig_ni = go.Figure()

    # Add traces Ni
    fig_ni.add_trace(go.Scatter(x=df.index, y=df['mral_ni'], mode='lines+markers', name='Ni',line=dict(color=colors['grade'])))
    fig_ni.add_trace(go.Scatter(x=df.index, y=df['oreas_ni'], mode='lines', name='Certificate',line=dict(color=colors['certificate'], width=2)))
    fig_ni.add_trace(go.Scatter(x=df.index, y=df['plus_ni_10'], mode='lines', name='Dil +10%',line=dict(color=colors['dil_10'])))
    fig_ni.add_trace(go.Scatter(x=df.index, y=df['plus_ni_5'], mode='lines', name='Dil +5%',line=dict(color=colors['dil_5'])))
    fig_ni.add_trace(go.Scatter(x=df.index, y=df['min_ni_10'], mode='lines', name='Dil -10%',line=dict(color=colors['dil_10'])))
    fig_ni.add_trace(go.Scatter(x=df.index, y=df['min_ni_5'], mode='lines', name='Dil -5%',line=dict(color=colors['dil_5'])))

    # Add mean trace
    fig_ni.add_trace(go.Scatter(
        x=[0, jml_row - 1],
        y=[avg_ni, avg_ni],
        mode='lines',
        name='Mean',
        # line=dict(color='#ffb100'),  dash="dot", width=3)
        line=dict(color=colors['mean'], dash="dot", width=3)
    ))

    fig_ni.update_layout(
        title=f'MRAL - Ni Analysis of {TypeCrm}',
        title_font=dict(size=18),
        margin=dict(l=40, r=25, t=35, b=25), 
        title_x=0.5,
        legend=dict(orientation="h"),
        hovermode='closest',
        plot_bgcolor='rgba(201,201,201,0.08)',  # Mengatur background plot 
        height=265, 
        xaxis=dict(
            overlaying='x',
            anchor='y',
            showticklabels=False,
        ),  
    )

    # Create Plotly figure - Fe
    fig_fe = go.Figure()

    # Add traces Fe
    fig_fe.add_trace(go.Scatter(x=df.index, y=df['mral_fe'], mode='lines+markers', name='Fe',line=dict(color=colors['grade'])))
    fig_fe.add_trace(go.Scatter(x=df.index, y=df['oreas_fe'], mode='lines', name='Certificate',line=dict(color=colors['certificate'])))
    fig_fe.add_trace(go.Scatter(x=df.index, y=df['plus_fe_10'], mode='lines', name='Dil +10%',line=dict(color=colors['dil_10'])))
    fig_fe.add_trace(go.Scatter(x=df.index, y=df['plus_fe_5'], mode='lines', name='Dil +5%',line=dict(color=colors['dil_5'])))
    fig_fe.add_trace(go.Scatter(x=df.index, y=df['min_fe_10'], mode='lines', name='Dil -10%',line=dict(color=colors['dil_10'])))
    fig_fe.add_trace(go.Scatter(x=df.index, y=df['min_fe_5'], mode='lines', name='Dil -5%',line=dict(color=colors['dil_5'])))

    # Add mean trace
    fig_fe.add_trace(go.Scatter(
        x=[0, jml_row - 1],
        y=[avg_fe, avg_fe],
        mode='lines',
        name='Mean',
        line=dict(color=colors['mean'],  dash="dot", width=3)
    ))

    fig_fe.update_layout(
        margin=dict(l=40, r=25, t=35, b=25), 
        title=f'MRAL - Fe Analysis of {TypeCrm}',
        title_font=dict(size=18),
        title_x=0.5,
        legend=dict(orientation="h"),
        hovermode='closest',
        plot_bgcolor='rgba(201,201,201,0.08)',  
        height=265,
        xaxis=dict(
            overlaying='x',  
            anchor='y',
            showticklabels=False,
        ),   
    )

    # Create Plotly figure - MgO
    fig_mgo = go.Figure()

    # Add traces Fe
    fig_mgo.add_trace(go.Scatter(x=df.index, y=df['mral_mgo'], mode='lines+markers', name='MgO',line=dict(color=colors['grade'])))
    fig_mgo.add_trace(go.Scatter(x=df.index, y=df['oreas_mgo'], mode='lines', name='Certificate',line=dict(color=colors['certificate'])))
    fig_mgo.add_trace(go.Scatter(x=df.index, y=df['plus_mgo_10'], mode='lines', name='Dil +10%',line=dict(color=colors['dil_10'])))
    fig_mgo.add_trace(go.Scatter(x=df.index, y=df['plus_mgo_5'], mode='lines', name='Dil +5%',line=dict(color=colors['dil_5'])))
    fig_mgo.add_trace(go.Scatter(x=df.index, y=df['min_mgo_10'], mode='lines', name='Dil -10%',line=dict(color=colors['dil_10'])))
    fig_mgo.add_trace(go.Scatter(x=df.index, y=df['min_mgo_5'], mode='lines', name='Dil -5%',line=dict(color=colors['dil_5'])))

    # Add mean trace
    fig_mgo.add_trace(go.Scatter(
        x=[0, jml_row - 1],
        y=[avg_mgo, avg_mgo],
        mode='lines',
        name='Mean',
        line=dict(color=colors['mean'],  dash="dot", width=3)
    ))

    fig_mgo.update_layout(
        title       =f'MRAL - MgO Analysis of {TypeCrm}',
        title_font  =dict(size=18),
        title_x     =0.5,
        margin      =dict(l=40, r=25, t=35, b=25), 
        legend      =dict(orientation="h"),
        hovermode   ='closest',
        plot_bgcolor='rgba(201,201,201,0.08)',  
        height      =265,
        xaxis=dict(
            overlaying='x',  # Menggunakan sumbu y yang sama dengan overlay
            anchor='y',
            showticklabels=False,
        ),
    )

    # Create Plotly figure - SiO2
    fig_sio2 = go.Figure()

    # Add traces Fe
    fig_sio2.add_trace(go.Scatter(x=df.index, y=df['mral_sio2'], mode='lines+markers', name='SiO2',line=dict(color=colors['grade'])))
    fig_sio2.add_trace(go.Scatter(x=df.index, y=df['oreas_sio2'], mode='lines', name='Certificate',line=dict(color=colors['certificate'])))
    fig_sio2.add_trace(go.Scatter(x=df.index, y=df['plus_sio2_10'], mode='lines', name='Dil +10%',line=dict(color=colors['dil_10'])))
    fig_sio2.add_trace(go.Scatter(x=df.index, y=df['plus_sio2_5'], mode='lines', name='Dil +5%',line=dict(color=colors['dil_5'])))
    fig_sio2.add_trace(go.Scatter(x=df.index, y=df['min_sio2_10'], mode='lines', name='Dil -10%',line=dict(color=colors['dil_10'])))
    fig_sio2.add_trace(go.Scatter(x=df.index, y=df['min_sio2_5'], mode='lines', name='Dil -5%',line=dict(color=colors['dil_5'])))

    # Add mean trace
    fig_sio2.add_trace(go.Scatter(
        x=[0, jml_row - 1],
        y=[avg_sio2, avg_sio2],
        mode='lines',
        name='Mean',
        line=dict(color=colors['mean'],  dash="dot", width=3)
    ))

    fig_sio2.update_layout(
        title       = f'MARL - SiO2 Analysis of {TypeCrm}',
        title_font  = dict(size=18),
        title_x     = 0.5,
        margin      = dict(l=40, r=25, t=35, b=25), 
        legend      = dict(orientation="h"),
        hovermode   = 'closest',
        plot_bgcolor= 'rgba(201,201,201,0.08)',  
        height      = 265,
        xaxis=dict(
            overlaying='x',  # Menggunakan sumbu y yang sama dengan overlay
            anchor='y',
            showticklabels=False,
        ),
    )

    
    # Convert figure to HTML
    plot_ni     = pio.to_html(fig_ni, full_html=False)
    plot_fe     = pio.to_html(fig_fe, full_html=False)
    plot_mgo    = pio.to_html(fig_mgo, full_html=False)
    plot_sio2   = pio.to_html(fig_sio2, full_html=False)

    # Create response data
    response_data = {
        # 'oreas_name': oreas_name,
        'jml_row'     :jml_row,
        'acceptedNi'  :true_ni,'errorNi':false_ni,'plotNi':plot_ni,  # Include plot HTML
        'acceptedFe'  :true_fe,'errorFe':false_fe,'plotFe':plot_fe,  # Include plot HTML
        'acceptedMgo' :true_mgo,'errorMgo':false_mgo,'plotMgo':plot_mgo,  # Include plot HTML
        'acceptedSio2':true_sio2,'errorSio2':false_sio2,'plotSio2':plot_sio2  # Include plot HTML
    }

    return JsonResponse(response_data)

#List Data Tables:
class sampleCrmMral(View):

    def post(self, request):
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)
    
    def _datatables(self, request):
        datatables          = request.POST
        draw                = int(datatables.get('draw'))
        start               = int(datatables.get('start'))
        length              = int(datatables.get('length'))
        search_value        = datatables.get('search[value]')
        order_column_index  = int(datatables.get('order[0][column]'))
        order_dir           = datatables.get('order[0][dir]')

        sql_query = """
            SELECT * FROM oreas_diff_abs_mral
        """

        params = []

      
        from_date      = request.POST.get('from_date')
        to_date        = request.POST.get('to_date')
        filterTypeCrm  = request.POST.get('filterTypeCrm')


        if from_date and to_date:
            sql_query += " WHERE release_date BETWEEN %s AND %s"
            params.extend([from_date, to_date])

        if filterTypeCrm:
            sql_query += " AND oreas_name = %s"
            params.extend([filterTypeCrm])    

        if search_value:
            sql_query += " AND (sample_number LIKE %s OR oreas_name LIKE %s)"
            params.extend([f"%{search_value}%", f"%{search_value}%"])
    
        if order_dir == 'desc':
            sql_query += f" ORDER BY {order_column_index} DESC"
        else:
            sql_query += f" ORDER BY {order_column_index} ASC"

        with connections['sqms_db'].cursor() as cursor:
            cursor.execute(sql_query, params)
            columns = [col[0] for col in cursor.description]
            sql_data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        total_records = len(sql_data)

        paginator   = Paginator(sql_data, length)
        total_pages = paginator.num_pages

        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = list(object_list)
        

        return {
            'draw'           : draw,
            'recordsTotal'   : total_records,
            'recordsFiltered': total_records,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }
    
@csrf_exempt
def export_crm_mral(request):
    from_date  = request.GET.get('from_date')
    to_date    = request.GET.get('to_date')
    filterTypeCrm  = request.GET.get('filterTypeCrm')

    # Query SQL
    sql_query = """
        SELECT
         oreas_name,oreas_ni ,oreas_co,oreas_fe2o3,oreas_fe,oreas_mgo
        ,oreas_sio2,sampling_deskripsi,sample_id,release_date,mral_ni
        ,mral_co,mral_fe2o3,mral_fe,mral_mgo,mral_sio2,diff_ni,diff_co,diff_fe2o3,diff_fe,diff_mgo,diff_sio2 
        FROM oreas_diff_abs_mral
    """
    params = []

    if from_date and to_date:
        sql_query += " WHERE release_date BETWEEN %s AND %s"
        params.extend([from_date, to_date])

    if filterTypeCrm:
            sql_query += " AND oreas_name = %s"
            params.extend([filterTypeCrm])    

    # Eksekusi query SQL dengan handling error jika data kosong
    with connections['sqms_db'].cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()
        
        # Cek apakah data kosong
        if not rows:
            return HttpResponse("No data available for export.", content_type="text/plain")
        
        columns = [col[0] for col in cursor.description]  # Ambil nama kolom dari hasil query

    # Inisialisasi Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data CRM - MRAL'

    # Header Kolom
    header = [
        'No', 'Name', 'Ni [Oreas]', 'Co [Oreas]', 'Fe2O3 [Oreas]', 'Fe [Oreas]', 'MgO [Oreas]', 'SiO2 [Oreas]', 'Deskripsi',
        'Sample Id', 'Realese','Ni', 'Co','Fe2O3','Fe', 'Mgo','SiO2','Diff [Ni]', 'Diff [Co]', 'Diff [Fe2O3]','Diff [Fe]', 'Diff [MgO]', 'Diff [SiO2]'
    ]

    # Cek kesesuaian jumlah header dengan data dari SQL
    expected_columns = len(header) - 1  # Kurangi 1 karena ada kolom "No"
    actual_columns = len(columns)

    if expected_columns != actual_columns:
        return HttpResponse(f"Column mismatch: expected {expected_columns}, got {actual_columns}.", content_type="text/plain")

    # Menulis header ke Excel
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    # Menulis data ke Excel
    for row_num, row in enumerate(rows, start=2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)  # Kolom "No"
        for col_num, cell_value in enumerate(row, start=2):  # Mulai dari kolom ke-2
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Menyesuaikan lebar kolom berdasarkan panjang data
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(column_title), max((len(str(row[col_num - 2])) for row in rows if row[col_num - 2] is not None), default=0))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Menyiapkan response untuk download file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Data CRM (MRAL).xlsx"'
    workbook.save(response)

    return response