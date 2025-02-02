# 
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.assay_mral_model import AssayMral
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from ....utils.permissions import get_dynamic_permissions


@login_required
def assay_mral_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    # Cek permission
    permissions = get_dynamic_permissions(request.user)
    context = {
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
        'permissions': permissions,
    }
    return render(request, 'admin-mgoqa/list-assay-mral.html',context)



class Assay_Mral(View):
    def post(self, request):
        # Ambil semua data invoice yang valid
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = AssayMral.objects.all()

        if search:
            data = data.filter(
                Q(job_number__icontains=search) |
                Q(sample_id__icontains=search) |
                Q(ni__icontains=search) 
            )
       

        # Filter berdasarkan parameter dari request
        startDate       = request.POST.get('startDate')
        endDate         = request.POST.get('endDate')


        if startDate and endDate:
            data = data.filter(release_date__range=[startDate, endDate])

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "release_mral": item.release_mral.strftime("%Y-%m-%d %H:%M:%S"),
                "job_number": item.job_number,
                "sample_id": item.sample_id,
                "ni": item.ni,
                "co": item.co,
                "fe2o3": item.fe2o3,
                "fe": item.fe,
                "mgo": item.mgo,
                "sio2": item.sio2
            } for item in object_list
        ]

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': total_records_filtered,
            'data': data,
            'start': start,
            'length': length,
            'totalPages': total_pages,
        }
    
@login_required
@csrf_exempt       
def export_data_mral(request):
    startDate = request.GET.get('startDate')
    endDate = request.GET.get('endDate')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data Ore'

    # Write header row
    header = [
        'No', 
        'Date', 
        'Job Number', 
        'Sample Id',
        'Ni',
        'Co',
        'Fe2O3',
        'Fe',
        'MgO',
        'SiO2'
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # List kolom yang ingin diambil
    columns = [
        'release_mral', 
        'job_number',
        'sample_id',
        'ni',
        'co',
        'fe2o3',
        'fe',
        'mgo',
        'sio2'
    ]

    queryset = AssayMral.objects.all().values_list(*columns)
    
    if startDate and endDate:
        queryset = queryset.filter(release_date__range=[startDate, endDate])

    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if isinstance(cell_value, datetime):  # Periksa jika nilai adalah datetime
                cell_value = cell_value.replace(tzinfo=None)  # Hapus zona waktu
            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data_mral.xlsx"'
    workbook.save(response)

    return response
