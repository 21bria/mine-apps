from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.sample_production_model import SampleProductions
from ....models.samples_data_view_model import SamplesView
from ....models.sample_method_details import SampleMethodDetail
from ....models.materials_model import Material
from ....models.source_model import SourceMinesDumping,SourceMinesDome
from ....models.selling_code_model import SellingCode
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from ....utils.utils import clean_string
from ....utils.permissions import get_dynamic_permissions

class SamplesDetails(View):

    def post(self, request):
        # Ambil semua data invoice yang valid
        data_sample = self._datatables(request)
        return JsonResponse(data_sample, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = SamplesView.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(sampling_area__icontains=search) |
                Q(sampling_point__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(batch_code__icontains=search) |
                Q(sample_number__icontains=search) |
                Q(no_sample__icontains=search)
            )
       

        # Filter berdasarkan parameter dari request
        from_date       = request.POST.get('from_date')
        to_date         = request.POST.get('to_date')
        materialFilter  = request.POST.get('materialFilter')
        typeFilter      = request.POST.get('typeFilter')
        areaFilter      = request.POST.get('areaFilter')
        pointFilter     = request.POST.get('pointFilter')
        factoriesFilter = request.POST.get('factoriesFilter')
        productFilter   = request.POST.get('productFilter')

        if from_date and to_date:
            data = data.filter(tgl_sample__range=[from_date, to_date])

        if typeFilter:
            data = data.filter(type_sample=typeFilter)

        if materialFilter:
            data = data.filter(nama_material=materialFilter)

        if areaFilter:
            data = data.filter(area_sampling=areaFilter)

        if pointFilter:
            data = data.filter(point_sampling=pointFilter)

        if factoriesFilter:
            data = data.filter(factory_stock=factoriesFilter)

        if productFilter:
            data = data.filter(product_code=productFilter)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"             : item.id,
                "tgl_sample"     : item.tgl_sample,
                "shift"          : item.shift,
                "type_sample"    : item.type_sample,
                "sample_method"  : item.sample_method,
                "sampling_area"  : item.sampling_area,
                "sampling_point" : item.sampling_point,
                "nama_material"  : item.nama_material,
                "batch_code"     : item.batch_code,
                "increments"     : item.increments,
                "size"           : item.size,
                "sample_weight"  : item.sample_weight,
                "sample_number"  : item.sample_number,
                "remark"         : item.remark,
                "primer_raw"     : item.primer_raw,
                "duplicate_raw"  : item.duplicate_raw,
                "to_its"         : item.to_its,
                "sampling_deskripsi": item.sampling_deskripsi,
                "no_sample"       : item.no_sample,
                "username"        : item.username
                
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }
    
@csrf_exempt      
def export_samples_data(request):
    # Lakukan filter data sesuai parameter yang diterima dari permintaan
    from_date       = request.GET.get('from_date')
    to_date         = request.GET.get('to_date')
    materialFilter  = request.GET.get('materialFilter')
    typeFilter      = request.GET.get('typeFilter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Export Data Ore'

    # Write header row
    header = [
        'No', 
        'Date',
        'Week',
        'Month',
        'Year',
        'Shift',
        'Type',
        'Method',
        'Material',
        'Sampling Area',
        'Sampling Point',
        'Batch',
        'Increments',
        'Sample Id',
        'Size',
        'Sample weight',
        'Primer raw',
        'Duplicate raw',
        'Sampling Description',
        'Factory',
        'Product Code'
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'tgl_sample', 
        'minggu', 
        'bulan',
        'tahun',
        'shift',
        'type_sample',
        'sample_method',
        'nama_material',
        'sampling_area',
        'sampling_point',
        'batch_code',
        'increments',
        'sample_number',
        'size',
        'sample_weight',
        'primer_raw',
        'duplicate_raw',
        'sampling_deskripsi',
        'factory_stock',
        'product_code'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = SamplesView.objects.all().values_list(*columns)
    
    if from_date and to_date:
            queryset = queryset.filter(tgl_sample__range=[from_date, to_date])

    if materialFilter:
            queryset = queryset.filter(nama_material=materialFilter)

    if typeFilter:
            queryset = queryset.filter(type_sample=typeFilter)        


    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Samples-data.xlsx"'
    workbook.save(response)

    return response

@login_required
def getIdSample(request, id):
    allowed_groups = ['superadmin','data-control','admin-mgoqa']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    if request.method == 'GET':
        try:
            items = SampleProductions.objects.get(id=id)
            dumping_point = None
            dome_point    = None
            nama_material = None
            code_name     = None

            if items.sampling_area:
                dumping = SourceMinesDumping.objects.filter(id=items.sampling_area).first()
                if dumping:
                    dumping_point = dumping.dumping_point

            if items.sampling_point:
                dome = SourceMinesDome.objects.filter(id=items.sampling_point).first()
                if dome:
                    dome_point = dome.pile_id

            if items.id_material:
                material = Material.objects.filter(id=items.id_material).first()
                if material:
                    nama_material = material.nama_material

            if items.product_code:
                product = SellingCode.objects.filter(id=items.product_code).first()
                if product:
                    code_name = product.product_code

            data = {
                'id'            : items.id,
                'tgl_sample'    : items.tgl_sample, 
                'shift'         : clean_string(items.shift),
                'id_type_sample': items.id_type_sample,
                'id_method'     : items.id_method,
                'id_material'   : items.id_material,
                'nama_material' : clean_string(nama_material),
                'sampling_area' : items.sampling_area,
                'dumping_point' : clean_string(dumping_point),
                'sampling_point': items.sampling_point,
                'dome_point'    : clean_string(dome_point),
                'product_code'  : items.product_code,
                'code_name'     : clean_string(code_name),
                'discharge_area': clean_string(items.discharge_area),
                'batch_code'    : clean_string(items.batch_code),
                'increments'    : items.increments,
                'sample_weight' : items.sample_weight,
                'sample_number' : clean_string(items.sample_number),
                'primer_raw'    : items.primer_raw,
                'duplicate_raw' : items.duplicate_raw,
                'to_its'        : items.to_its,
                'unit_truck'    : clean_string(items.unit_truck),
                'type'          : clean_string(items.type),
                'sampling_deskripsi': clean_string(items.sampling_deskripsi),
                'gc_expect'     : items.gc_expect,
                'remark'        : clean_string(items.remark),
                
            }
            return JsonResponse(data)
        except SampleProductions.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

def get_methodSample(request):
    type_id = request.GET.get('type_id')

    # Periksa apakah type_id ada dan valid
    if not type_id:
        return JsonResponse({'error': 'Type ID is required'}, status=400)

    # Pastikan untuk mengkonversi type_id ke integer
    try:
        type_id = int(type_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid Type ID'}, status=400)

    # Ambil method berdasarkan type_id
    method_sample = SampleMethodDetail.objects.filter(type_id=type_id).values('id', 'sample_method')
    
    if not method_sample:
        return JsonResponse([], safe=False)  # Kembalikan list kosong jika tidak ada kota

    return JsonResponse(list(method_sample), safe=False)
    
@login_required
def deleteSample(request):
    allowed_groups = ['superadmin']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if job_id:
            # Lakukan penghapusan berdasarkan ID di sini
            data = SampleProductions.objects.get(id=int(job_id))
            data.delete()
            return JsonResponse({'status': 'deleted'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
def samples_data_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    
    # Ambil permissions dinamis dari database
    permissions = get_dynamic_permissions(request.user)
    context = {
        'day_date'   : today.strftime('%Y-%m-%d'),
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
        'permissions': permissions,
    }
    
    return render(request, 'admin-mgoqa/production-samples/list-sample.html', context)
